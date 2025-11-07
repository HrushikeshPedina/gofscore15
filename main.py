# golf_peoria_modern.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import plotly.express as px

st.set_page_config(page_title="Peoria Stableford — Modern Dashboard", layout="wide")

# -------------------------------
# Styling (dark navy theme)
# -------------------------------
st.markdown(
    """
    <style>
    .stApp, .main {
      background-color: #051826;
      color: #FFFFFF;
    }
    .block-container {
      background-color: rgba(5, 24, 38, 0.98);
      color: #FFFFFF;
      padding: 1.25rem;
      border-radius: 12px;
    }
    h1, h2, h3 { color: #FFD966 !important; }
    .stButton>button, .stDownloadButton>button {
      background-color: #FFD966 !important;
      color: #052235 !important;
      font-weight: 600;
      border-radius: 8px;
    }
    .stCheckbox>label { color: #FFFFFF !important; }
    </style>
    """, unsafe_allow_html=True
)

# -------------------------------
# Helper functions
# -------------------------------
def stableford_points_from_net(par, net_score):
    diff = net_score - par
    if diff >= 2:
        return 0
    elif diff == 1:
        return 1
    elif diff == 0:
        return 2
    elif diff == -1:
        return 3
    elif diff == -2:
        return 4
    else:
        return 5

def compute_peoria_allowance_float(pars, scores, ref_holes):
    """Compute Peoria allowance as float: sum(score-par) on ref holes * 1.5"""
    adjustments = []
    for h in ref_holes:
        idx = h - 1
        adjustments.append(scores[idx] - pars[idx])
    return sum(adjustments) * 1.5

def allocate_strokes(handicap_int, stroke_indexes):
    """
    Distribute integer strokes across 15 holes based on stroke index.
    stroke_indexes is length-15 list with values 1..15 (1 hardest).
    Allocate base = H // 15 to all holes, then remainder to holes with smallest stroke_index.
    Returns list length 15 with per-hole allocated strokes.
    """
    H = max(0, int(round(handicap_int)))
    base = H // 15
    rem = H % 15
    alloc = [base] * 15
    if rem > 0:
        # give +1 to holes with stroke_index 1..rem
        for i, si in enumerate(stroke_indexes):
            if si <= rem:
                alloc[i] += 1
    return alloc

# -------------------------------
# Processing workbook
# -------------------------------
def process_workbook_bytes(file_bytes, selected_peoria_holes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Read hole rows 2..16 (15 holes)
    hole_nums = [ws.cell(row=r, column=1).value for r in range(2, 17)]
    pars = [ws.cell(row=r, column=2).value for r in range(2, 17)]
    stroke_idx = [ws.cell(row=r, column=3).value for r in range(2, 17)]

    # Validate stroke index
    if not all(isinstance(si, int) and 1 <= si <= 15 for si in stroke_idx):
        raise ValueError("Stroke_Index values must be integers 1..15 for each hole (rows 2..16).")

    # Player names from header row 1, columns 4..
    player_cols = list(range(4, ws.max_column + 1))
    player_names = [ws.cell(row=1, column=c).value or f"Player_{i+1}" for i, c in enumerate(player_cols)]

    # Build long-form rows: player x hole
    rows = []
    players_detail = []  # will store per-player detail dicts
    for i_col, col in enumerate(player_cols):
        name = player_names[i_col]
        # read scores (rows 2..16)
        scores = [ws.cell(row=r, column=col).value for r in range(2, 17)]
        scores = [int(s) for s in scores]  # assume valid ints

        # compute peoria allowance float
        peoria_float = compute_peoria_allowance_float(pars, scores, selected_peoria_holes)
        peoria_int = int(round(peoria_float)) if peoria_float >= 0 else 0

        # allocate strokes
        strokes_alloc = allocate_strokes(peoria_int, stroke_idx)

        # per-hole net and stableford
        net_scores = [scores[h] - strokes_alloc[h] for h in range(15)]
        stableford_per_hole = [stableford_points_from_net(pars[h], net_scores[h]) for h in range(15)]

        gross = sum(scores)
        net_total = sum(net_scores)
        total_stableford = sum(stableford_per_hole)

        # append detail rows
        for h in range(15):
            rows.append({
                "Player": name,
                "Hole": hole_nums[h],
                "Par": pars[h],
                "Stroke_Index": stroke_idx[h],
                "Gross_Score": scores[h],
                "Strokes_Allocated": strokes_alloc[h],
                "Net_Score": net_scores[h],
                "Stableford": stableford_per_hole[h]
            })

        players_detail.append({
            "Player": name,
            "Scores": scores,
            "Strokes_Allocated": strokes_alloc,
            "Net_Per_Hole": net_scores,
            "Stableford_Per_Hole": stableford_per_hole,
            "Peoria_Float": round(peoria_float, 2),
            "Peoria_Int": peoria_int,
            "Gross": gross,
            "Net_Total": net_total,
            "Total_Stableford": total_stableford
        })

    # create DataFrame long form
    df_long = pd.DataFrame(rows)

    # summary
    summary_rows = []
    for pdict in players_detail:
        summary_rows.append({
            "Player": pdict["Player"],
            "Gross": pdict["Gross"],
            "Peoria_Int": pdict["Peoria_Int"],
            "Net_Total": pdict["Net_Total"],
            "Stableford": pdict["Total_Stableford"]
        })
    df_summary = pd.DataFrame(summary_rows)

    # Rankings
    # Best gross (ties)
    min_gross = df_summary["Gross"].min()
    best_gross_players = df_summary[df_summary["Gross"] == min_gross]

    # Best gross per group of 4
    group_best = []
    for i in range(0, len(df_summary), 4):
        grp = df_summary.iloc[i:i+4]
        if grp.empty:
            continue
        m = grp["Gross"].min()
        winners = grp[grp["Gross"] == m]
        group_best.append(winners)

    top_stableford = df_summary.sort_values("Stableford", ascending=False).head(10)
    top_5_net = df_summary.sort_values("Net_Total", ascending=True).head(5)

    # Build Excel report (bytes)
    wb_out = Workbook()
    ws_det = wb_out.active
    ws_det.title = "PerPlayerDetails"

    # Header row for details
    header = ["Player", *[f"H{h}" for h in range(1,16)]]
    # Scores
    ws_det.append(["Players & Gross Scores"] + [""] * 15)
    for p in players_detail:
        ws_det.append([p["Player"]] + p["Scores"])
    ws_det.append([])
    # Strokes allocated
    ws_det.append(["Strokes Allocated"] + [""] * 15)
    for p in players_detail:
        ws_det.append([p["Player"]] + p["Strokes_Allocated"])
    ws_det.append([])
    # Net per hole
    ws_det.append(["Net Scores"] + [""] * 15)
    for p in players_detail:
        ws_det.append([p["Player"]] + p["Net_Per_Hole"])
    ws_det.append([])
    # Stableford per hole
    ws_det.append(["Stableford (Net)"] + [""] * 15)
    for p in players_detail:
        ws_det.append([p["Player"]] + p["Stableford_Per_Hole"])
    # Summary sheet
    ws_sum = wb_out.create_sheet("Summary")
    ws_sum.append(["Player", "Gross", "Peoria_Int", "Net_Total", "Stableford"])
    for _, r in df_summary.iterrows():
        ws_sum.append([r["Player"], r["Gross"], r["Peoria_Int"], r["Net_Total"], r["Stableford"]])
    ws_sum.append([])
    ws_sum.append(["Best Gross (ties)"])
    for _, r in best_gross_players.iterrows():
        ws_sum.append([r["Player"], r["Gross"]])
    ws_sum.append([])
    ws_sum.append(["Top 10 Stableford"])
    for i, r in enumerate(top_stableford.itertuples(index=False), start=1):
        ws_sum.append([i, r.Player, r.Stableford])
    ws_sum.append([])
    ws_sum.append(["Top 5 Net"])
    for i, r in enumerate(top_5_net.itertuples(index=False), start=1):
        ws_sum.append([i, r.Player, r.Net_Total])

    out_bytes = BytesIO()
    wb_out.save(out_bytes)
    out_bytes.seek(0)

    return {
        "df_long": df_long,
        "df_summary": df_summary,
        "players_detail": players_detail,
        "best_gross_players": best_gross_players,
        "group_best": group_best,
        "top_stableford": top_stableford,
        "top_5_net": top_5_net,
        "excel_bytes": out_bytes
    }

# -------------------------------
# Streamlit UI layout
# -------------------------------
st.title("⛳ Peoria Stableford — Modern Dashboard (Option B)")

st.markdown("Upload a 15-hole scorecard Excel with columns: **Hole | Par | Stroke_Index | Player1 | Player2 | ...** (Hole rows: 1–15 in rows 2–16).")

uploaded = st.file_uploader("Upload score_card_new.xlsx", type=["xlsx"])
if not uploaded:
    st.info("Please upload the scorecard Excel file to begin.")
    st.stop()

# preview uploaded file top rows
try:
    wb_preview = openpyxl.load_workbook(BytesIO(uploaded.read()), data_only=True)
    ws_preview = wb_preview.active
    headers = [ws_preview.cell(row=1, column=c).value or f"Col{c}" for c in range(1, ws_preview.max_column+1)]
    preview_rows = []
    for r in range(1, 17):  # header + 15 holes
        preview_rows.append([ws_preview.cell(row=r, column=c).value for c in range(1, ws_preview.max_column+1)])
    df_preview = pd.DataFrame(preview_rows, columns=headers)
    st.subheader("📋 Uploaded File Preview (first 16 rows)")
    st.dataframe(df_preview, use_container_width=True)
except Exception as e:
    st.error(f"Failed to preview uploaded file: {e}")
    st.stop()

st.markdown("### 🎯 Select exactly 10 Peoria holes (these are the 10 holes used to calculate Peoria allowance)")
selected_holes = st.multiselect("Choose 10 holes", options=list(range(1,16)))
if len(selected_holes) != 10:
    st.warning("Please select exactly 10 holes to continue.")
    st.stop()

# Process
with st.spinner("Calculating..."):
    try:
        uploaded.seek(0)
        result = process_workbook_bytes(uploaded.read(), selected_holes)
    except Exception as e:
        st.error(f"Processing failed: {e}")
        st.stop()

# Display summary and long table
st.success("✅ Calculations complete!")

# two-column top summary + leaderboards
colA, colB = st.columns([2,1])

with colA:
    st.subheader("🏁 Player Summary")
    df_sum_disp = result["df_summary"].sort_values("Stableford", ascending=False).reset_index(drop=True)
    st.dataframe(df_sum_disp.style.format({"Gross": "{:.0f}", "Net_Total":"{:.0f}", "Stableford":"{:.0f}"}), use_container_width=True)

with colB:
    st.subheader("🏆 Leaderboards")
    st.markdown("**Best Gross (ties)**")
    for _, r in result["best_gross_players"].iterrows():
        st.write(f"- **{r.Player}** — Gross {r.Gross}")
    st.markdown("**Top 10 Stableford**")
    st.table(result["top_stableford"].reset_index(drop=True))
    st.markdown("**Top 5 Net**")
    st.table(result["top_5_net"].reset_index(drop=True))

# Unified hole-by-hole table (Option B)
st.subheader("📖 Hole-by-Hole (Unified Table)")
df_long_display = result["df_long"].copy()
# reorder columns for clarity
df_long_display = df_long_display[["Player","Hole","Par","Stroke_Index","Gross_Score","Strokes_Allocated","Net_Score","Stableford"]]
st.dataframe(df_long_display, use_container_width=True)

# Charts: Gross / Stableford / Net (side-by-side)
st.subheader("📊 Visual Analysis")
df_chart = result["df_summary"].copy()

c1, c2, c3 = st.columns(3)
with c1:
    fig_gross = px.bar(df_chart, x="Player", y="Gross", text="Gross", color="Gross", color_continuous_scale="Blues")
    fig_gross.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
    st.plotly_chart(fig_gross, use_container_width=True)
with c2:
    fig_net = px.bar(df_chart, x="Player", y="Net_Total", text="Net_Total", color="Net_Total", color_continuous_scale="Cividis")
    fig_net.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
    st.plotly_chart(fig_net, use_container_width=True)
with c3:
    fig_sf = px.bar(df_chart, x="Player", y="Stableford", text="Stableford", color="Stableford", color_continuous_scale="Viridis")
    fig_sf.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
    st.plotly_chart(fig_sf, use_container_width=True)

# Hole-by-hole aggregate: average Stableford per hole across players
st.subheader("📈 Hole-by-Hole Aggregate (Average Stableford per Hole)")
df_hole_agg = df_long_display.groupby("Hole").agg(
    Avg_Stableford=("Stableford","mean"),
    Avg_Net=("Net_Score","mean"),
    Avg_Gross=("Gross_Score","mean")
).reset_index()
fig_hole = px.line(df_hole_agg, x="Hole", y=["Avg_Stableford","Avg_Net","Avg_Gross"], markers=True,
                   labels={"value":"Average", "variable":"Metric"}, title="Averages per Hole")
fig_hole.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
st.plotly_chart(fig_hole, use_container_width=True)

# Expandable per-player detail (still provide expanders, though main table exists)
st.subheader("🔎 Expandable Player Details")
for p in result["players_detail"]:
    with st.expander(p["Player"], expanded=False):
        df_p = pd.DataFrame({
            "Hole": list(range(1,16)),
            "Par": p["Scores"] and result["df_long"].loc[result["df_long"]["Player"]==p["Player"], "Par"].unique().tolist() or [],
            "Gross": p["Scores"],
            "Strokes_Allocated": p["Strokes_Allocated"],
            "Net": p["Net_Per_Hole"],
            "Stableford": p["Stableford_Per_Hole"]
        })
        st.dataframe(df_p, use_container_width=True)
        # per-player charts
        fig_line = px.line(df_p, x="Hole", y=["Gross","Net"], markers=True, title=f"{p['Player']} — Gross vs Net")
        fig_line.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
        st.plotly_chart(fig_line, use_container_width=True)
        fig_bar = px.bar(df_p, x="Hole", y="Stableford", text="Stableford", title=f"{p['Player']} — Stableford per Hole")
        fig_bar.update_layout(template="plotly_dark", paper_bgcolor="#051826", plot_bgcolor="#051826", title_font_color="#FFD966", font_color="white")
        st.plotly_chart(fig_bar, use_container_width=True)

# Best gross per group (display)
st.subheader("🥇 Best Gross by Group of 4 (ties allowed)")
for gi, grp in enumerate(result["group_best"], start=1):
    st.write(f"Group {gi}: " + ", ".join(f"{row.Player} ({row.Gross})" for _, row in grp.iterrows()))

# Download full results excel
st.markdown("### 📥 Download full Excel report")
st.download_button("Download Excel Report", data=result["excel_bytes"].getvalue(), file_name="peoria_stableford_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("---")
st.caption("Peoria Stableford calculations: allowance computed on selected 10 holes as sum(score-par)*1.5; integer allowance distributed to holes by Stroke_Index (1 = hardest). Stableford uses net score (gross - strokes allocated).")
